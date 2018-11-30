#!/usr/bin/python
#encoding=utf-8

__author__ = "小白龙"

from openpyxl import Workbook  #写
from openpyxl import load_workbook  #读
import datetime
import urllib.request

def writeExcl():
    wb = Workbook()
    sheet1 = wb.active
    #设置sheet名字
    sheet1.title = "第一个工作簿"
    sheet1['A1'] = "书名"
    #在尾部添加一行
    sheet1.append([1, 2, 4])
    sheet1['B1'] = "时间"
    sheet1.append([datetime.datetime.now()])
    #保存文件
    wb.save("写入excel示例.xlsx")

def readExcel(path):
    print("读取Excel：",path)
    wb = load_workbook(path)
    sheets = wb.get_sheet_names()
    print(type(sheets))
    print("工作簿：",sheets)
    # 获取要读取的sheet
    ws = wb.get_sheet_by_name(sheets[0])

    #读取单元格的三种方式：
    print("读取某一单元格的值", ws['A1'].value)
    # print("读取某一单元格的值", ws.cell('A1'))
    print("读取某一单元格的值", ws.cell(row=1,column=1))

    print("打印表头：")
    for cell in ('A1','B1','C1','D1'):
        if cell == 'D1':
            print(ws[cell].value)
        else:
            print(ws[cell].value,", ",end="")
    for cell in ('A2','B2','C2','D2'):
        if cell == 'D2':
            print(ws[cell].value)
        else:
            print(ws[cell].value,", ",end="")
    for cell in ('A3','B3','C3','D3'):
        if cell == 'D3':
            print(ws[cell].value)
        else:
            print(ws[cell].value,", ",end="")
    for cell in ('A4','B4','C4','D4'):
        if cell == 'D4':
            print(ws[cell].value)
        else:
            print(ws[cell].value,", ",end="")
    #获取连续单元格，只是获取单元格，不是读取其值
    cell_range = ws['A1':'A2']
    # for cell in cell_range:
    #     print(ws[cell])
    print(cell_range)
    #获取所有单元格
    cl = ws.get_cell_collection()
    print (type(cl))
    print("根据行列值获取单元格:")
    print(ws.cell(row=1,column=2).value)


def searchBook():
    url = "https://book.douban.com/j/subject_suggest?q=python"
    response = urllib.request.urlopen(url)
    book_str = response.read().decode()
    print(book_str)
    book_dict = eval(book_str)
    for book in book_dict:
        print(book["title"]," ",book["author_name"]," ",book["year"])
    print("将书籍信息写入excel：")
    wb = Workbook()
    ws = wb.active
    ws.append(["书名","作者","发布时间","链接","图片链接","书号"])
    for book in book_dict:
        ws.append([book["title"],book["author_name"],book["year"],book["url"],book["pic"],book["id"]])
    ws.append(["本次记录书籍共计：",len(book_dict)])
    print("本次记录书籍共计：",len(book_dict))

    wb.save("豆瓣搜书结果.xlsx")


if __name__ == "__main__":
    writeExcl()
    readExcel("读取Excel示例.xlsx")
    searchBook()